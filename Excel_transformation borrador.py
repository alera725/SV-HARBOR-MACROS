# -*- coding: utf-8 -*-
"""
Created on Wed Sep  1 09:57:50 2021

@author: alejandro.gutierrez
"""

import os
import shutil
#os.chdir('C:\\Users\\alejandro.gutierrez\\OneDrive - Carlin Group - CA Fortune\\Documents\\ALEJANDRO RAMOS GTZ\\GIT\\RITE INSIGHT BOT') # relative path: scripts dir is under Lab

from openpyxl import Workbook
from openpyxl import load_workbook
import pandas as pd 
import numpy as np
from datetime import date, timedelta
import datetime

#%% DAY

#Set paths 
import_path = 'C:\\Users\\alejandro.gutierrez\\OneDrive - Carlin Group - CA Fortune\\Documents\\KROGER SELENIUM\\SUPER VALU\\MONTHLY_DC'
export_path = 'C:\\SUPERVALU SQL'

#Fecha de hoy
today = date.today()

#Encontrar el utlimo dia sabado
idx = (today.weekday() + 1) % 7
sat = today - datetime.timedelta(7+idx-6)  

d1 = sat.strftime("%d/%m/%Y") #Last Saturday complete date

sun = sat - datetime.timedelta(days=6)
d2 = sun.strftime("%d/%m/%Y") #Previous Sunday before that Saturday 

#Week number 
t1 = sun
week_number = t1.strftime("%U")

#REVISAR SI ESTA SEMANA TOCO CAMBIO DE MES
#Revisamos el primer dia el mes y el ultimo dia el mes, si hubo cambio vamos a ver en que dia es el ultimo o primero del nuevo mes 
#Unique values de la cadena de todos los months de los dias de esa semana ==2 si cambio el mes, otro no
delta = sat - sun       # as timedelta
month_list = []
year_list = []

for i in range(delta.days + 1):
    day = sun + timedelta(days=i)
    month_list.append(day.strftime("%d/%m/%Y")[3:5]) #list of the month of each day
    year_list.append(day.strftime("%d/%m/%Y")[6:10])


#month_list = ['09', '09', '09', '10', '10', '10', '10'] #ELIMINAR

month_list_drop_duplicates = list(set(month_list))        
num_unique_months = len(month_list_drop_duplicates) #If we have more than 2 months we have to do something different than if we have 1 unique month

year_list_drop_duplicates = list(set(year_list))
num_unique_years = len(year_list_drop_duplicates)

#Leer los files dependiendo su hubo cambio de periodo anual, mes o nada 

if (num_unique_months==2):
    # Leer los archivos separados
    data_sv1 = pd.read_csv(import_path + '\\FIRST SUPER VALU MONTLHY DC week %s.csv'%week_number, skiprows=5) #leer week 53 'FIRST SUPER VALU MONTLHY DC week 53
    data_sv2 = pd.read_csv(import_path + '\\SECOND SUPER VALU MONTLHY DC week %s.csv'%week_number, skiprows=5) #leer week 1 'SECOND SUPER VALU MONTLHY DC week 1'
    
    for i in range(2):
        # Renombrar columnas con espacios en blanco
        data_sv = data_sv.rename(columns={'Invoice Week': 'Invoice_Week', 'Corp Code': 'Corp_Code', 'Order Type': 'Order_Type', 'Vendor#': 'Vendorno', 'Product Group': 'Product_Group', 'Class_Group': 'Class_Group', 'Supporting DC Item': 'Supporting_DC_Item'})
        
        # Cambiar el tipo de dato de dos columnas a numericas (no fue necesario pero ver VWPOUNDS A INT64 Y SVNETWT A INT64)
        data_sv['SVNetWt'] = data_sv['SVNetWt'].astype('int64')
        # Es diferente el proceso ya que VWPounds era float64 pero si tenia valores decimales en la base no como SVNetWt
        data_sv['VWPounds'] = data_sv['VWPounds'].fillna(0).astype(np.int64, errors='ignore')
        
        # Exportamos el archivo a la ubicacion deseada 
        data_sv.to_csv(export_path + '\\DataSV.csv', index = False)
        
else:
    # Leer el archivo 
    data_sv = pd.read_csv(import_path + '\\SUPER VALU MONTLHY DC week %s.csv'%week_number, skiprows=5) #'SUPER VALU MONTLHY DC 25-Aug-2021 14Hr 44Min'
    
    # Renombrar columnas con espacios en blanco
    data_sv = data_sv.rename(columns={'Invoice Week': 'Invoice_Week', 'Corp Code': 'Corp_Code', 'Order Type': 'Order_Type', 'Vendor#': 'Vendorno', 'Product Group': 'Product_Group', 'Class_Group': 'Class_Group', 'Supporting DC Item': 'Supporting_DC_Item'})
    
    # Cambiar el tipo de dato de dos columnas a numericas (no fue necesario pero ver VWPOUNDS A INT64 Y SVNETWT A INT64)
    data_sv['SVNetWt'] = data_sv['SVNetWt'].astype('int64')
    # Es diferente el proceso ya que VWPounds era float64 pero si tenia valores decimales en la base no como SVNetWt
    data_sv['VWPounds'] = data_sv['VWPounds'].fillna(0).astype(np.int64, errors='ignore')
    
    # Exportamos el archivo a la ubicacion deseada 
    data_sv.to_csv(export_path + '\\DataSV.csv', index = False)


#%% PANDAS
data_sv = pd.read_csv(import_path + '\\SUPER VALU MONTLHY DC week %s.csv'%week_number, skiprows=5)

# Renombrar columnas con espacios en blanco
data_sv = data_sv.rename(columns={'Invoice Week': 'Invoice_Week', 'Corp Code': 'Corp_Code', 'Order Type': 'Order_Type', 'Vendor#': 'Vendorno', 'Product Group': 'Product_Group', 'Class_Group': 'Class_Group', 'Supporting DC Item': 'Supporting_DC_Item'})
#print(data_sv.columns)

#Cambiar el tipo de dato de dos columnas a numericas (no fue necesario pero ver VWPOUNDS A INT64 Y SVNETWT A INT64)
data_sv.dtypes

data_sv['SVNetWt'] = data_sv['SVNetWt'].astype('int64')
#Es diferente el proceso ya que VWPounds era float64 pero si tenia valores decimales en la base no como SVNetWt
data_sv['VWPounds'] = data_sv['VWPounds'].fillna(0).astype(np.int64, errors='ignore')


data_sv.to_csv('C:\\Users\\alejandro.gutierrez\\OneDrive - Carlin Group - CA Fortune\\Desktop\\DataSV.csv', index = False)


#Ver un archivo de ejemplo de SV anteriores
datasv2 = pd.read_csv('C:\\Users\\alejandro.gutierrez\\OneDrive - Carlin Group - CA Fortune\\Desktop\\DataSV Segundo.csv')
datasv2.dtypes

#%% EXCEL NO FUE NECESARIO
workbook = load_workbook(filename = import_path + '\\SUPER VALU MONTLHY DC week 34.csv')
workbook.sheetnames #Get the sheet names 

#Activate the sheet to work in  
sheet = workbook.active

#Write what you want in a specific cell
sheet["C1"] = 'Hola'

#Other way
cell = sheet["A1"]
cell.value = 'Hello'

#Delete rows 
sheet.delete_rows(idx=1, amount=4) #Delete first 4 rows 

# Convert column to number formatting 
col = sheet.column_dimensions['U']
col.number_format = '0'

#other way
number_format = '0'
for col in (5,7):
    for row in range(2,sheet.max_row+1):
        sheet.cell(row = row, column = col).style = number_format

#Get values of Excel Sheet to into in DF
data = sheet.values
#Set the 1st row as the column headers for DF
cols = next(data)
data = list(data)

df = pd.DataFrame(data, columns = cols)

df.to_csv('.csv')

#Save the spreadsheet
workbook.save(filename = ".xlsx")

